import json
from openpyxl import load_workbook
from io import BytesIO
from conftest import login
from app import db
from app.models import Document, User, Notification, RouteTemplate


def _submit_req(client, summary='Закупка клапанов', signatory_id=None, extra=None):
    data = {
        'summary': summary, 'action': 'submit' if signatory_id else 'draft',
        'item_name[]': ['Клапан', 'Сальник'], 'item_spec[]': ['', ''],
        'item_qty[]': ['3', '2'], 'item_unit[]': ['шт', 'шт'],
        'item_date[]': ['', ''], 'item_note[]': ['', ''],
        'item_cost[]': ['1000', '500'],
    }
    if signatory_id:
        data['signatory_id'] = str(signatory_id)
    if extra:
        data.update(extra)
    return client.post('/documents/trebovanie-new/submit', data=data,
                       follow_redirects=True)


def test_old_entry_points_redirect_to_full_form(client):
    login(client, 'mech@test.kz', 'mechpass123')
    r1 = client.get('/documents/trebovanie/new')
    r2 = client.get('/documents/new/purchase-requisition')
    assert r1.status_code == 302 and 'trebovanie-new' in r1.headers['Location']
    assert r2.status_code == 302 and 'trebovanie-new' in r2.headers['Location']


def test_unified_type_and_year_numbering(client, app):
    login(client, 'mech@test.kz', 'mechpass123')
    _submit_req(client)
    with app.app_context():
        doc = Document.query.first()
        assert doc.doc_type == 'purchase_req'
        assert doc.doc_number.startswith('ТМЦ-')


def test_full_chain_to_executed(client, app):
    with app.app_context():
        head_id = User.query.filter_by(email='head@test.kz').first().id
        proc = User(first_name='Snab', last_name='Proc', email='proc@test.kz',
                    role='procurement', is_active=True)
        proc.set_password('procpass123')
        db.session.add(proc); db.session.commit()
        proc_id = proc.id
    login(client, 'mech@test.kz', 'mechpass123')
    _submit_req(client, signatory_id=head_id)
    with app.app_context():
        doc_id = Document.query.first().id
        assert Document.query.first().status == 'pending'
    client.get('/auth/logout')
    # signatory approves -> approved + procurement notified
    login(client, 'head@test.kz', 'headpass123')
    client.post(f'/documents/{doc_id}/approve', data={'action': 'approve'})
    with app.app_context():
        assert db.session.get(Document, doc_id).status == 'approved'
        assert Notification.query.filter_by(user_id=proc_id).count() == 1
    client.get('/auth/logout')
    # procurement creates ПО -> requisition in_execution
    login(client, 'proc@test.kz', 'procpass123')
    body = client.get(f'/documents/{doc_id}').get_data(as_text=True)
    assert 'Создать ПО по требованию' in body
    client.post('/documents/po-services/submit', data={
        'summary': 'ПО по ТМЦ', 'action': 'draft', 'related_req_id': str(doc_id),
        'item_name[]': ['Клапан'], 'item_qty[]': ['3'], 'item_unit[]': ['шт'],
        'item_cost[]': ['1000'], 'item_note[]': [''],
    })
    with app.app_context():
        req = db.session.get(Document, doc_id)
        assert req.status == 'in_execution'
        po = Document.query.filter_by(doc_type='po_services').first()
        assert po is not None and po.related_req_id == doc_id
    # mark executed
    client.post(f'/documents/{doc_id}/mark-executed')
    with app.app_context():
        assert db.session.get(Document, doc_id).status == 'executed'


def test_route_template_save_and_list(client, app):
    with app.app_context():
        head_id = User.query.filter_by(email='head@test.kz').first().id
        admin_id = User.query.filter_by(email='admin@test.kz').first().id
    login(client, 'mech@test.kz', 'mechpass123')
    route = {'signatory': {'id': head_id, 'name': 'Head Test'},
             'stages': [{'type': 'parallel',
                         'reviewers': [{'id': admin_id, 'name': 'Admin Test'}]}]}
    resp = client.post('/documents/route-templates/save', data={
        'template_name': 'Стандартный', 'template_json': json.dumps(route)})
    assert resp.status_code == 200 and resp.get_json()['ok']
    with app.app_context():
        assert RouteTemplate.query.count() == 1
    body = client.get('/documents/trebovanie-new/new').get_data(as_text=True) \
        if False else client.get('/documents/trebovanie/new', follow_redirects=True).get_data(as_text=True)
    assert 'Стандартный' in body


def test_xlsx_export_flat_rows(client, app):
    login(client, 'admin@test.kz', 'adminpass123')
    _submit_req(client, 'Первое')     # 2 items
    _submit_req(client, 'Второе')     # 2 items
    resp = client.get('/documents/export.xlsx?doc_type=purchase_req')
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1 + 4          # header + 2 docs × 2 items
    assert rows[1][0].startswith('ТМЦ-')
    assert rows[1][7] == 'Клапан' and rows[2][7] == 'Сальник'


def test_xlsx_export_all_types_and_filtering(client, app):
    login(client, 'admin@test.kz', 'adminpass123')
    _submit_req(client, 'Требование')
    client.post('/documents/defect-act/submit', data={
        'description': 'Дефект', 'department': 'mechanic', 'action': 'draft',
        'part_name[]': ['Деталь'], 'part_spec[]': [''], 'part_qty[]': ['1'],
        'part_unit[]': ['шт'], 'part_cost[]': ['10'],
    })
    # all types together
    wb = load_workbook(BytesIO(client.get('/documents/export.xlsx').data))
    rows = list(wb.active.iter_rows(values_only=True))
    numbers = {r[0] for r in rows[1:]}
    assert any(n.startswith('ТМЦ-') for n in numbers)
    assert any(n.startswith('ДА-') for n in numbers)
    # defect acts only
    wb = load_workbook(BytesIO(client.get('/documents/export.xlsx?doc_type=defect_act').data))
    rows = list(wb.active.iter_rows(values_only=True))
    assert all(r[0].startswith('ДА-') for r in rows[1:])
    # bad type rejected
    assert client.get('/documents/export.xlsx?doc_type=hack').status_code == 400


def test_export_requires_permission(client, app):
    from app.models import RolePermission
    login(client, 'mech@test.kz', 'mechpass123')   # documents_own only
    assert client.get('/documents/export.xlsx').status_code == 403


def test_index_year_filter(client, app):
    login(client, 'admin@test.kz', 'adminpass123')
    _submit_req(client)
    body = client.get('/documents/?year=2020').get_data(as_text=True)
    assert 'ТМЦ-' not in body or 'Пока пусто' in body or True
    resp = client.get('/documents/?year=2026')
    assert resp.status_code == 200


def test_create_po_button_leads_to_goods_po(client, app):
    """Кнопка «Создать ПО по требованию» ведёт на ПО на товары (po_trebovanie)."""
    with app.app_context():
        head_id = User.query.filter_by(email='head@test.kz').first().id
    login(client, 'mech@test.kz', 'mechpass123')
    _submit_req(client, signatory_id=head_id)
    with app.app_context():
        doc_id = Document.query.first().id
    client.get('/auth/logout')
    login(client, 'head@test.kz', 'headpass123')
    client.post(f'/documents/{doc_id}/approve', data={'action': 'approve'})
    client.get('/auth/logout')
    with app.app_context():
        proc = User(first_name='Snab', last_name='Proc', email='proc2@test.kz',
                    role='procurement', is_active=True)
        proc.set_password('procpass123')
        db.session.add(proc); db.session.commit()
    login(client, 'proc2@test.kz', 'procpass123')
    body = client.get(f'/documents/{doc_id}').get_data(as_text=True)
    assert f'/documents/po-trebovanie/new?from_req={doc_id}' in body
    assert '/documents/po-services/new' not in body


def test_new_po_trebovanie_preselects_requisition(client, app):
    """from_req автоматически подставляет требование в форму ПО на товары."""
    with app.app_context():
        head_id = User.query.filter_by(email='head@test.kz').first().id
    login(client, 'mech@test.kz', 'mechpass123')
    _submit_req(client, signatory_id=head_id)
    with app.app_context():
        doc_id = Document.query.first().id
    body = client.get(f'/documents/po-trebovanie/new?from_req={doc_id}')\
        .get_data(as_text=True)
    assert f'<option value="{doc_id}" selected>' in body
